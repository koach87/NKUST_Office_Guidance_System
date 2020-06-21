from tkinter import *
from tkinter import messagebox
import pandas as pd
import webbrowser
from openpyxl import Workbook, load_workbook
from os import path
from datetime import datetime

class Gui_For_Office:

    def __init__(self,master):
        self.master = master
        self.initUI()

    def callback(self,url):
        webbrowser.open_new(url)

    def add_data_to_history(self, identity, number, name, dept, question):

        now = datetime.now()
        time = now.strftime("%Y/%m/%d %H:%M")

        wb = load_workbook(filename = self.history_file_name)
        ws = wb.active
        data = [time, identity, number, name, dept, question]
        row_add = ws.max_row+1
        for i, x in enumerate(data):
            print(x, i)
            ws.cell(column=i+1, row=row_add, value = x)
        
        wb.save(self.history_file_name)
        
        

    def to_start_page(self):
        self.btn_cancel.pack_forget()
        self.std_id.delete(0, 'end')
        self.questions_page.place_forget()
        self.tchr_page.place_forget()
        self.std_page.place_forget()
        self.start_page.place(relx=0.5, rely=0.5, anchor=CENTER)

    def to_questions_page(self):
        self.btn_cancel.pack(anchor="n", side =LEFT, padx = 10, pady = 10)
        self.tchr_page.place_forget()
        self.start_page.place_forget()
        self.questions_page.place(relx=0.5, rely=0.5, anchor=CENTER)

    def to_student_page(self):
        self.start_page.place_forget()
        self.tchr_page.place_forget()
        self.btn_cancel.pack(anchor="n", side =LEFT, padx = 10, pady = 10)
        self.std_page.place(relx=0.5, rely=0.5, anchor=CENTER)

    def to_dept_page(self):
        self.btn_cancel.pack(anchor="n", side =LEFT, padx = 10, pady = 10)
        self.start_page.place_forget()
        self.tchr_page.place(relx=0.5, rely=0.5, anchor=CENTER)
        

# ================================================================================
    # click action
    def click_tchr(self):
        self.to_dept_page()

    def click_tchr_OK(self):
        pass

    def click_std(self):
        self.to_student_page()

    def click_std_OK(self):
        # print(self.xls_stdlist.get(self.std_id.get()))
        if(self.xls_stdlist.get(self.std_id.get())):
            self.std_page.place_forget()
            self.to_questions_page()
        else:
            messagebox.showerror('找不到該學號','請重新輸入，或至點擊下方連結查詢學號')
        
        


    def return_tchr_dept(self, dept_index):
        self.xls_dept_seat = pd.read_excel("GUI_services.xlsx", sheet_name= 1 )
      

        messagebox.showinfo('國立高雄科技大學進推處教務組','請老師至 {} 號櫃台，由承辦人員為您服務。'
                            .format(self.xls_dept_seat.values[dept_index][1]))

        # import net
        # try:
        #     net.host("{}老師 t1 t2".format(self.xls_dept_seat.values[dept_index][0]), 22222)
        # except ConnectionRefusedError:
        #     print('cannot link')      

        self.add_data_to_history('導師','','',self.dept_list[dept_index],'')

        self.to_start_page()
            
        

    def return_question(self,question_index):
        num = self.std_id.get() 
        name = self.xls_stdlist.get(num)[0]
        dept = self.xls_stdlist.get(num)[1]
        tel = self.xls_stdlist.get(num)[2]
        dept_num = self.xls_stdlist.get(num)[3]
        question =  self.xls_services.values[question_index][0]
        question_remarks = self.xls_services.values[question_index][1]

        print(num, dept, name, question)
                
        messagebox.showinfo("國立高雄科技大學進推處教務組","{} \n請至 {} 號櫃台\n備註:\n{}"
                                .format(question, dept_num, question_remarks))
        print("分機:{}".format(tel))

        self.to_start_page()
        self.add_data_to_history('學生', num, name, dept, question)

        # import net
        # try:
        #     net.host("學生 {}系學生{}\n學號:{}\n辦理{}".format(dept, name, num, question), 22222)
        # except ConnectionRefusedError:
        #     print('cannot link') 
    
    def initUI(self):


        # set UI 
        self.master.iconbitmap("nkust.ico")
        self.master.state("zoom")
        self.master.title("Office Guidance System")
 
# ================================================================================ cancel frame
        
        cancel_frame = Frame(self.master, bg = "green").pack()
        self.btn_cancel = Button(cancel_frame, text = "取消", width = 10, font = ("微軟正黑體",25),command = self.to_start_page)

# ================================================================================ start page 
        
        self.start_page = Frame(self.master, bg="gray")
        start_page_middle = Frame(self.start_page,bg = "white")

        # Label of notice
        Label(start_page_middle, text = "請選擇身份", font = ("微軟正黑體",30)).pack(pady = 20)

        # middle widget
        self.btn_std = Button(start_page_middle,text = "學生", width = 10, height = 5, font = ("微軟正黑體",30), command = self.click_std)
        self.btn_std.pack(side = LEFT,padx = 50)

        # middle widget
        self.btn_tchr = Button(start_page_middle,text = "導師", width = 10, height = 5, font = ("微軟正黑體",30), command = self.click_tchr)
        self.btn_tchr.pack(side = LEFT,padx = 50)

        # middle frame setting
        start_page_middle.pack()
        # start_page_middle.place(relx=0.5, rely=0.5, anchor=CENTER)

# ================================================================================ student page

        # get stdlist
        self.xls_stdlist = pd.read_excel("GUI_stdlist.xlsx", sheet_name= 0 )
        self.xls_stdlist = self.xls_stdlist.astype(str).set_index('std_num').T.to_dict('list')
        
        self.std_page = Frame(self.master)

        Label(self.std_page,text = "請輸入學號", font = ("微軟正黑體",25)).pack(pady = 10)

        self.std_id = Entry(self.std_page, width = 50, justify = CENTER, font = ("微軟正黑體",25))
        self.std_id.pack(pady = 10)

        # middle widget
        std_btn_OK = Button(self.std_page,text = "確定", font = ("微軟正黑體",25), width = 25,command = self.click_std_OK)
        std_btn_OK.pack(pady = 10)

        self.look_for_id = Label(self.std_page, text = "查詢學號", font = ("微軟正黑體",15), fg = "blue", cursor = "hand2")
        self.look_for_id.pack(pady = 10)
        self.look_for_id.bind("<Button-1>", lambda e: self.callback("https://webap.nkust.edu.tw/nkust/system/getuid.jsp?kind=2"))

        

# ================================================================================ qeustions page
        # get services
        self.xls_services = pd.read_excel("GUI_services.xlsx", sheet_name= 0 )

        self.questions_page = Frame(self.master, bg = "black")

        questions_label = Label(self.questions_page, font = ("微軟正黑體",30), text = "請點選欲辦理項目")
        questions_label.pack()

        service_buttons_frame = Frame(self.questions_page)
        # service_buttons_frame.place(relx=0.5, rely=0.5, anchor=CENTER)
        service_buttons_frame.pack()
        services_cnt = 12

        for i in range(services_cnt//3):
            for j in range(3):
                cnt = i*3+j
                b = Button(service_buttons_frame,text = self.xls_services.values[cnt][0] ,width = 30 ,height = 2 ,font =("微軟正黑體",24),
                            command = lambda x = cnt : self.return_question(x))
                b.grid(row = i, column = j, padx = 40, pady = 40)


# ================================================================================ teacher's page
        
        self.dept_list = ''
        with open('GUI_dept.txt', encoding= 'UTF-8') as f:
            self.dept_list = f.read().split(' ')

        self.tchr_page = Frame(self.master)
        Label(self.tchr_page, font = ("微軟正黑體",30), text = "請選擇所屬科系").pack()
        dept_buttons_frame = Frame(self.tchr_page)
        dept_buttons_frame.pack()
        dept_cnt = 20
        row = 4
        for i in range(dept_cnt//row):
            for j in range(row):
                cnt = i*row+j
                b = Button(dept_buttons_frame,text = self.dept_list[cnt] ,width = 15 ,height = 1 ,font =("微軟正黑體",24),
                        command = lambda x = cnt : self.return_tchr_dept(x))
                b.grid(row = i, column = j, padx = 40, pady = 40)


        
        # # nkust image
        # im = PhotoImage(file = "nkust.gif")
        # print(im)
        # aa = Canvas(self.master)
        # aa.pack()
        # aa.create_image(20,20, image = im)


        # self.questions_page.pack()
        # self.std_page.pack()
        

        month = datetime.today().strftime('%m')
        self.history_file_name = '進推處教務組{}月承辦紀錄.xlsx'.format(month)
        print(self.history_file_name)
        # print(path.exists(history_file_name))
        if not (path.exists(self.history_file_name)):
            wb = Workbook()
            ws = wb.active
            data = ['時間', '身分', '學號', '姓名', '科系', '事項']
            for i, x in enumerate(data):
                print(x, i)
                ws.cell(column=i+1, row=1, value = x)
            
            wb.save(self.history_file_name)

        self.to_start_page()
    

if __name__ == "__main__":
    a = Gui_For_Office(Tk())
    a.master.mainloop()