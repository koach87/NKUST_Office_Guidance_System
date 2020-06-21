
from openpyxl import Workbook, load_workbook
from os import path
from datetime import datetime

month = datetime.today().strftime('%m')
history_file_name = '進推處教務組{}月承辦紀錄.xlsx'.format(month)
data_title = ['時間', '身分', '學號', '姓名', '科系', '事項']

def check_data_exist():
    # print(history_file_name)
    # print(path.exists(history_file_name))

    if not (path.exists(history_file_name)):
        wb = Workbook()
        ws = wb.active
        for i, x in enumerate(data_title):
            # print(x, i)
            ws.cell(column=i+1, row=1, value = x)
        
        wb.save(history_file_name)

def add_data_to_history(identity, number, name, dept, question):

    now = datetime.now()
    time = now.strftime("%Y/%m/%d %H:%M")

    wb = load_workbook(filename = history_file_name)
    ws = wb.active
    data = [time, identity, number, name, dept, question]
    row_add = ws.max_row+1
    for i, x in enumerate(data):
        # print(x, i)
        ws.cell(column=i+1, row=row_add, value = x)
    
    wb.save(history_file_name)

